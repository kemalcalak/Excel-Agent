"""Local .xlsx editing for files stored in Google Drive.

Design decisions:
- .xlsx files in Drive are NEVER converted to Google Sheets format.
- Flow: download from Drive -> edit locally with openpyxl -> upload back to
  the same Drive ID. The Drive ID, sharing settings and links are preserved.
- Native Google Sheets files are handled by GoogleSheetsTools, not this module.
"""

from __future__ import annotations

import io
import json
import re
from pathlib import Path
from typing import Any, List, Optional, Union

from agno.tools import Toolkit
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import range_boundaries

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SHEETS_MIME = "application/vnd.google-apps.spreadsheet"
FOLDER_MIME = "application/vnd.google-apps.folder"

# Local working directory for downloaded .xlsx files.
WORKDIR = Path("excel_workdir")


def load_credentials(
    creds_path: str = "credentials.json",
    token_path: str = "token.json",
) -> Credentials:
    """Run the OAuth flow and persist the token.

    The same credentials object is shared by GoogleDriveTools,
    GoogleSheetsTools and ExcelTools so the user only consents once.
    """
    token_file = Path(token_path)
    creds: Optional[Credentials] = None

    if token_file.exists():
        creds = Credentials.from_authorized_user_file(str(token_file), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(creds_path, SCOPES)
            creds = flow.run_local_server(port=0)
        token_file.write_text(creds.to_json())

    return creds


def _local_path(file_id: str) -> Path:
    WORKDIR.mkdir(exist_ok=True)
    return WORKDIR / f"{file_id}.xlsx"


class ExcelTools(Toolkit):
    """Toolkit for editing Drive-resident .xlsx files without format conversion.

    Workflow:
      1) download_excel(file_id)            -> fetch the file locally
      2) list_sheet_names / read_excel_range -> inspect contents
      3) update_excel_cell / update_excel_range / append_excel_row /
         create_excel_sheet_tab             -> edit locally
      4) upload_excel(file_id)              -> push back to the same Drive ID
                                                (still .xlsx)
    """

    def __init__(self, creds: Credentials):
        self.creds = creds
        self.drive = build("drive", "v3", credentials=creds)
        super().__init__(
            name="excel_tools",
            tools=[
                self.download_excel,
                self.upload_excel,
                self.list_sheet_names,
                self.read_excel_range,
                self.update_excel_cell,
                self.update_excel_range,
                self.append_excel_row,
                self.create_excel_sheet_tab,
                # High-level operations
                self.find_and_replace_excel,
                self.find_cells_excel,
                self.delete_rows_excel,
                self.delete_rows_where,
                self.insert_rows_excel,
                self.filter_rows_excel,
                self.column_summary_excel,
                self.set_formula_excel,
                # Folder-level operations
                self.find_folder_by_name,
                self.list_excels_in_folder,
                self.bulk_find_replace_in_folder,
                # Drive file management
                self.move_drive_file_to_trash,
                self.rename_drive_file,
                self.copy_drive_file,
                # Sheet tab management (Drive .xlsx)
                self.rename_excel_sheet_tab,
                self.delete_excel_sheet_tab,
                # Column operations + sort (Drive .xlsx)
                self.sort_excel_by_column,
                self.delete_excel_columns,
                self.insert_excel_columns,
                # Analytical + export + creation (Drive .xlsx)
                self.describe_excel,
                self.export_drive_excel_to_csv,
                self.create_drive_xlsx_file,
            ],
        )

    # ---------- Drive transfer ----------

    def download_excel(self, file_id: str) -> str:
        """Download a Drive .xlsx file to `excel_workdir/{file_id}.xlsx`.

        Args:
            file_id: Drive file ID. The file MUST have the .xlsx mimeType;
                     do not call this for native Google Sheets files.

        Returns:
            JSON: {local_path, name} or an error message.
        """
        try:
            meta = self.drive.files().get(fileId=file_id, fields="name, mimeType").execute()
            mime = meta.get("mimeType")
            if mime != EXCEL_MIME:
                return (
                    f"File is not .xlsx (mimeType={mime}). "
                    "Use GoogleSheetsTools for native Google Sheets files; DO NOT convert."
                )

            request = self.drive.files().get_media(fileId=file_id)
            buf = io.BytesIO()
            downloader = MediaIoBaseDownload(buf, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()

            path = _local_path(file_id)
            path.write_bytes(buf.getvalue())
            return json.dumps(
                {"local_path": str(path), "name": meta["name"]}, ensure_ascii=False
            )
        except Exception as e:
            return f"Download error: {e}"

    def upload_excel(self, file_id: str) -> str:
        """Upload the edited local .xlsx back to the SAME Drive ID (overwrite).

        The format is preserved — the file stays as .xlsx and is not converted
        to Google Sheets. Drive ID, link and sharing settings are unchanged.

        Args:
            file_id: Target Drive file ID (same one used in download_excel).

        Returns:
            Success message or error.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            media = MediaIoBaseUpload(
                io.BytesIO(path.read_bytes()), mimetype=EXCEL_MIME, resumable=False
            )
            self.drive.files().update(fileId=file_id, media_body=media).execute()
            return f"Uploaded to Drive (xlsx format preserved): {file_id}"
        except Exception as e:
            return f"Upload error: {e}"

    # ---------- Local read ----------

    def list_sheet_names(self, file_id: str) -> str:
        """List sheet (tab) names inside the local .xlsx.

        Args:
            file_id: Drive ID of a file previously fetched with download_excel.

        Returns:
            JSON array of sheet names.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            wb = load_workbook(path, read_only=True, data_only=False)
            names = wb.sheetnames
            wb.close()
            return json.dumps(names, ensure_ascii=False)
        except Exception as e:
            return f"Read error: {e}"

    def read_excel_range(
        self,
        file_id: str,
        sheet_name: str,
        cell_range: Optional[str] = None,
        values_only: bool = True,
    ) -> str:
        """Read cells from a specific sheet.

        Args:
            file_id: Drive file ID (call download_excel first).
            sheet_name: Sheet/tab name, e.g. "Sayfa1".
            cell_range: Range like "A1:D20". If None, the full used range
                is returned.
            values_only: If True, only values; otherwise the textual cell repr.

        Returns:
            JSON 2D list.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
            target = ws[cell_range] if cell_range else ws.iter_rows(values_only=values_only)

            rows: List[List[Any]] = []
            if cell_range:
                # ws[range] returns a tuple for a single cell; normalise to 2D.
                if not isinstance(target, tuple):
                    target = (target,)
                for row in target:
                    if not isinstance(row, tuple):
                        row = (row,)
                    rows.append([c.value for c in row])
            else:
                for row in target:
                    rows.append(list(row) if isinstance(row, tuple) else [row])
            wb.close()
            return json.dumps(rows, ensure_ascii=False, default=str)
        except Exception as e:
            return f"Read error: {e}"

    # ---------- Local write ----------

    def update_excel_cell(
        self, file_id: str, sheet_name: str, cell: str, value: Any
    ) -> str:
        """Update a single cell.

        Args:
            file_id: Drive file ID.
            sheet_name: Sheet name.
            cell: Cell reference, e.g. "B5".
            value: Value to write. Formulas like "=SUM(A1:A10)" are accepted.

        Returns:
            Success message.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            wb[sheet_name][cell] = value
            wb.save(path)
            wb.close()
            return f"{sheet_name}!{cell} = {value!r} written (local). Don't forget upload_excel to push to Drive."
        except Exception as e:
            return f"Write error: {e}"

    def update_excel_range(
        self, file_id: str, sheet_name: str, start_cell: str, data: List[List[Any]]
    ) -> str:
        """Write a 2D block of values into a range.

        Args:
            file_id: Drive file ID.
            sheet_name: Sheet name.
            start_cell: Top-left anchor, e.g. "A1". Data flows right and down.
            data: 2D list; inner-list length defines the number of columns.

        Returns:
            Success message.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
            start_col, start_row, _, _ = range_boundaries(start_cell + ":" + start_cell)
            for r, row in enumerate(data):
                for c, val in enumerate(row):
                    ws.cell(row=start_row + r, column=start_col + c, value=val)
            wb.save(path)
            wb.close()
            end_cell = f"{get_column_letter(start_col + max((len(r) for r in data), default=1) - 1)}{start_row + len(data) - 1}"
            return f"{sheet_name}!{start_cell}:{end_cell} written (local). Don't forget upload_excel to push to Drive."
        except Exception as e:
            return f"Write error: {e}"

    def append_excel_row(
        self, file_id: str, sheet_name: str, row_values: List[Any]
    ) -> str:
        """Append a new row at the bottom of a sheet.

        Args:
            file_id: Drive file ID.
            sheet_name: Sheet name.
            row_values: List of cell values for the new row.

        Returns:
            Success message.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            wb[sheet_name].append(row_values)
            wb.save(path)
            wb.close()
            return f"Row appended to {sheet_name} (local). Don't forget upload_excel to push to Drive."
        except Exception as e:
            return f"Write error: {e}"

    def create_excel_sheet_tab(self, file_id: str, sheet_name: str) -> str:
        """Add a new sheet (tab) to the local .xlsx file.

        Args:
            file_id: Drive file ID.
            sheet_name: Name of the new sheet to create.

        Returns:
            Success message.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            wb = load_workbook(path)
            if sheet_name in wb.sheetnames:
                return f"Sheet already exists: {sheet_name}"
            wb.create_sheet(title=sheet_name)
            wb.save(path)
            wb.close()
            return f"New sheet added: {sheet_name} (local). Call upload_excel to push."
        except Exception as e:
            return f"Sheet creation error: {e}"

    # ---------- High-level operations ----------

    def find_and_replace_excel(
        self,
        file_id: str,
        find: str,
        replace: str,
        sheet_name: Optional[str] = None,
        column: Optional[str] = None,
        case_sensitive: bool = False,
        whole_cell: bool = False,
        use_regex: bool = False,
    ) -> str:
        """Bulk find-and-replace across the workbook or a single sheet.

        Args:
            file_id: Drive file ID.
            find: Substring or regex pattern to look for.
            replace: Replacement text.
            sheet_name: Restrict to a single sheet; None scans every sheet.
            column: Restrict to a single column letter (e.g. "C") or 1-based
                index as a string. None scans every column.
            case_sensitive: Case sensitivity for substring/regex matching.
            whole_cell: Require the cell value to equal `find` exactly
                (mutually exclusive with substring semantics).
            use_regex: Treat `find` as a regular expression.

        Returns:
            JSON: {replacements_per_sheet: {...}, total: N}.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            wb = load_workbook(path)
            sheets = [sheet_name] if sheet_name else wb.sheetnames
            for s in sheets:
                if s not in wb.sheetnames:
                    return f"Sheet not found: {s}. Available: {wb.sheetnames}"

            col_idx = _resolve_column(column) if column else None

            pattern: Optional[re.Pattern[str]] = None
            if use_regex:
                flags = 0 if case_sensitive else re.IGNORECASE
                pattern = re.compile(find, flags)

            counts: dict[str, int] = {}
            for s in sheets:
                ws = wb[s]
                n = 0
                for row in ws.iter_rows():
                    for cell in row:
                        if col_idx is not None and cell.column != col_idx:
                            continue
                        if cell.value is None:
                            continue
                        text = str(cell.value)

                        new_text: Optional[str] = None
                        if whole_cell:
                            if use_regex and pattern and pattern.fullmatch(text):
                                new_text = replace
                            elif not use_regex and (
                                (case_sensitive and text == find)
                                or (not case_sensitive and text.lower() == find.lower())
                            ):
                                new_text = replace
                        else:
                            if use_regex and pattern:
                                replaced = pattern.sub(replace, text)
                                if replaced != text:
                                    new_text = replaced
                            elif not use_regex:
                                if case_sensitive:
                                    if find in text:
                                        new_text = text.replace(find, replace)
                                else:
                                    if find.lower() in text.lower():
                                        new_text = re.sub(
                                            re.escape(find), replace, text, flags=re.IGNORECASE
                                        )

                        if new_text is not None:
                            cell.value = new_text
                            n += 1
                counts[s] = n

            wb.save(path)
            wb.close()
            total = sum(counts.values())
            return json.dumps(
                {
                    "replacements_per_sheet": counts,
                    "total": total,
                    "note": "Local file updated. Call upload_excel to push to Drive.",
                },
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Find/replace error: {e}"

    def find_cells_excel(
        self,
        file_id: str,
        sheet_name: str,
        pattern: str,
        case_sensitive: bool = False,
        whole_cell: bool = False,
        use_regex: bool = False,
        max_results: int = 200,
    ) -> str:
        """Locate cells whose value matches a pattern.

        Args:
            file_id: Drive file ID.
            sheet_name: Sheet to scan.
            pattern: Substring or regex.
            case_sensitive: Case sensitivity flag.
            whole_cell: Require exact match.
            use_regex: Interpret pattern as regex.
            max_results: Cap on matches returned.

        Returns:
            JSON list of {cell, value} entries.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]

            regex: Optional[re.Pattern[str]] = None
            if use_regex:
                flags = 0 if case_sensitive else re.IGNORECASE
                regex = re.compile(pattern, flags)
            needle = pattern if case_sensitive else pattern.lower()

            hits: list[dict[str, Any]] = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value)
                    cmp_text = text if case_sensitive else text.lower()
                    match = False
                    if use_regex and regex:
                        match = bool(regex.fullmatch(text) if whole_cell else regex.search(text))
                    else:
                        match = (cmp_text == needle) if whole_cell else (needle in cmp_text)
                    if match:
                        hits.append({"cell": cell.coordinate, "value": text})
                        if len(hits) >= max_results:
                            break
                if len(hits) >= max_results:
                    break
            wb.close()
            return json.dumps(
                {"matches": hits, "count": len(hits), "truncated": len(hits) >= max_results},
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Search error: {e}"

    def delete_rows_excel(
        self, file_id: str, sheet_name: str, start_row: int, count: int = 1
    ) -> str:
        """Delete a contiguous block of rows.

        Args:
            file_id: Drive file ID.
            sheet_name: Sheet name.
            start_row: 1-based row index to start deleting from.
            count: Number of rows to delete (default 1).

        Returns:
            Success message.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            wb[sheet_name].delete_rows(start_row, count)
            wb.save(path)
            wb.close()
            return f"{sheet_name}: {count} row(s) deleted starting at row {start_row}. Don't forget upload_excel."
        except Exception as e:
            return f"Delete error: {e}"

    def delete_rows_where(
        self,
        file_id: str,
        sheet_name: str,
        column: str,
        operator: str,
        value: Optional[Union[str, float, int]] = None,
        has_header: bool = True,
    ) -> str:
        """Delete every row where a column value matches a condition.

        Args:
            file_id: Drive file ID.
            sheet_name: Sheet name.
            column: Column letter ("B") or 1-based index as string ("2").
            operator: One of `==`, `!=`, `>`, `<`, `>=`, `<=`, `contains`,
                `not_contains`, `empty`, `not_empty`.
            value: Comparison value. Ignored for `empty`/`not_empty`.
            has_header: Skip the first row if True.

        Returns:
            JSON: {deleted, remaining}.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
            col_idx = _resolve_column(column)

            rows_to_delete: list[int] = []
            start = 2 if has_header else 1
            for row_idx in range(start, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if _match_condition(cell_value, operator, value):
                    rows_to_delete.append(row_idx)

            for r in reversed(rows_to_delete):
                ws.delete_rows(r, 1)

            wb.save(path)
            wb.close()
            return json.dumps(
                {
                    "deleted": len(rows_to_delete),
                    "remaining": ws.max_row,
                    "note": "Local file updated. Don't forget upload_excel.",
                },
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Conditional delete error: {e}"

    def insert_rows_excel(
        self, file_id: str, sheet_name: str, before_row: int, count: int = 1
    ) -> str:
        """Insert blank rows above the given 1-based row index.

        Args:
            file_id: Drive file ID.
            sheet_name: Sheet name.
            before_row: New rows are inserted starting at this index.
            count: Number of blank rows to insert.

        Returns:
            Success message.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            wb[sheet_name].insert_rows(before_row, count)
            wb.save(path)
            wb.close()
            return f"{sheet_name}: {count} blank row(s) inserted before row {before_row}. Don't forget upload_excel."
        except Exception as e:
            return f"Row insert error: {e}"

    def filter_rows_excel(
        self,
        file_id: str,
        sheet_name: str,
        column: str,
        operator: str,
        value: Optional[Union[str, float, int]] = None,
        has_header: bool = True,
        max_results: int = 200,
    ) -> str:
        """Return rows matching a column condition without modifying the file.

        Args:
            file_id: Drive file ID.
            sheet_name: Sheet name.
            column: Column letter or 1-based index as string.
            operator: Same operators as `delete_rows_where`.
            value: Comparison value.
            has_header: If True, header row is returned separately.
            max_results: Cap on matched rows.

        Returns:
            JSON: {header, rows, count, truncated}.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
            col_idx = _resolve_column(column)

            header: Optional[List[Any]] = None
            rows: list[list[Any]] = []
            start = 2 if has_header else 1
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if has_header and i == 1:
                    header = list(row)
                    continue
                if i < start:
                    continue
                cell_value = row[col_idx - 1] if col_idx - 1 < len(row) else None
                if _match_condition(cell_value, operator, value):
                    rows.append(list(row))
                    if len(rows) >= max_results:
                        break
            wb.close()
            return json.dumps(
                {
                    "header": header,
                    "rows": rows,
                    "count": len(rows),
                    "truncated": len(rows) >= max_results,
                },
                ensure_ascii=False,
                default=str,
            )
        except Exception as e:
            return f"Filter error: {e}"

    def column_summary_excel(
        self,
        file_id: str,
        sheet_name: str,
        column: str,
        has_header: bool = True,
    ) -> str:
        """Compute basic statistics for a column's numeric values.

        Args:
            file_id: Drive file ID.
            sheet_name: Sheet name.
            column: Column letter or 1-based index as string.
            has_header: Skip the first row when True.

        Returns:
            JSON: {count, numeric_count, sum, mean, min, max,
                   non_numeric_count, empty_count}.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
            col_idx = _resolve_column(column)

            nums: list[float] = []
            empty = 0
            non_numeric = 0
            total = 0
            start_iter = ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True)
            for i, row in enumerate(start_iter, start=1):
                if has_header and i == 1:
                    continue
                total += 1
                v = row[0]
                if v is None or (isinstance(v, str) and not v.strip()):
                    empty += 1
                    continue
                try:
                    nums.append(float(v))
                except (TypeError, ValueError):
                    non_numeric += 1
            wb.close()

            summary: dict[str, Any] = {
                "count": total,
                "numeric_count": len(nums),
                "non_numeric_count": non_numeric,
                "empty_count": empty,
            }
            if nums:
                summary.update(
                    sum=sum(nums),
                    mean=sum(nums) / len(nums),
                    min=min(nums),
                    max=max(nums),
                )
            return json.dumps(summary, ensure_ascii=False)
        except Exception as e:
            return f"Summary error: {e}"

    def set_formula_excel(
        self,
        file_id: str,
        sheet_name: str,
        target_column: str,
        formula_template: str,
        start_row: int,
        end_row: int,
    ) -> str:
        """Apply a templated formula to a column range, substituting `{row}`.

        Example template: `=C{row}*D{row}` written to E2..E100 yields
        `=C2*D2`, `=C3*D3`, ... per row.

        Args:
            file_id: Drive file ID.
            sheet_name: Sheet name.
            target_column: Destination column letter or 1-based index string.
            formula_template: Formula text; `{row}` is replaced by the row index.
            start_row: First row to write (1-based, inclusive).
            end_row: Last row to write (1-based, inclusive).

        Returns:
            Success message.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        if start_row > end_row:
            return "start_row cannot be greater than end_row."
        try:
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
            col_idx = _resolve_column(target_column)
            for r in range(start_row, end_row + 1):
                ws.cell(row=r, column=col_idx, value=formula_template.replace("{row}", str(r)))
            wb.save(path)
            wb.close()
            col_letter = get_column_letter(col_idx)
            return f"Formula written to {sheet_name}!{col_letter}{start_row}:{col_letter}{end_row}. Don't forget upload_excel."
        except Exception as e:
            return f"Formula write error: {e}"

    # ---------- Folder-level operations ----------

    def find_folder_by_name(self, name: str, max_results: int = 10) -> str:
        """Search Drive for folders whose name contains the given text.

        Args:
            name: Substring to search for (case-insensitive in Drive).
            max_results: Cap on results.

        Returns:
            JSON list of {id, name, modifiedTime}.
        """
        try:
            safe = name.replace("'", "\\'")
            query = (
                f"mimeType='{FOLDER_MIME}' and name contains '{safe}' "
                "and trashed=false"
            )
            result = (
                self.drive.files()
                .list(
                    q=query,
                    pageSize=max_results,
                    fields="files(id, name, modifiedTime)",
                    orderBy="modifiedTime desc",
                )
                .execute()
            )
            files = result.get("files", [])
            return json.dumps({"folders": files, "count": len(files)}, ensure_ascii=False)
        except Exception as e:
            return f"Folder search error: {e}"

    def list_excels_in_folder(
        self,
        folder_id: str,
        include_sheets: bool = False,
        recursive: bool = False,
        max_results: int = 200,
    ) -> str:
        """List spreadsheet files inside a Drive folder.

        Args:
            folder_id: Drive folder ID (obtain via find_folder_by_name).
            include_sheets: If True, also include native Google Sheets files
                alongside .xlsx files. Default False (Excel only).
            recursive: If True, traverse subfolders breadth-first.
            max_results: Hard cap across the whole walk.

        Returns:
            JSON: {files: [{id, name, mimeType, parent_path}], count}.
        """
        try:
            mime_clause = (
                f"(mimeType='{EXCEL_MIME}' or mimeType='{SHEETS_MIME}')"
                if include_sheets
                else f"mimeType='{EXCEL_MIME}'"
            )

            collected: list[dict[str, Any]] = []
            # Each entry: (folder_id, path_label)
            queue: list[tuple[str, str]] = [(folder_id, "")]
            seen: set[str] = set()

            while queue and len(collected) < max_results:
                current_id, path = queue.pop(0)
                if current_id in seen:
                    continue
                seen.add(current_id)

                files_query = f"'{current_id}' in parents and {mime_clause} and trashed=false"
                page_token: Optional[str] = None
                while True:
                    res = (
                        self.drive.files()
                        .list(
                            q=files_query,
                            pageSize=min(100, max_results - len(collected)),
                            fields="nextPageToken, files(id, name, mimeType, modifiedTime)",
                            orderBy="name",
                            pageToken=page_token,
                        )
                        .execute()
                    )
                    for f in res.get("files", []):
                        f["parent_path"] = path or "."
                        collected.append(f)
                        if len(collected) >= max_results:
                            break
                    page_token = res.get("nextPageToken")
                    if not page_token or len(collected) >= max_results:
                        break

                if recursive and len(collected) < max_results:
                    subq = (
                        f"'{current_id}' in parents and mimeType='{FOLDER_MIME}' "
                        "and trashed=false"
                    )
                    subres = (
                        self.drive.files()
                        .list(q=subq, fields="files(id, name)", pageSize=100)
                        .execute()
                    )
                    for sub in subres.get("files", []):
                        new_path = f"{path}/{sub['name']}" if path else sub["name"]
                        queue.append((sub["id"], new_path))

            return json.dumps(
                {"files": collected, "count": len(collected)}, ensure_ascii=False
            )
        except Exception as e:
            return f"Folder listing error: {e}"

    def bulk_find_replace_in_folder(
        self,
        folder_id: str,
        find: str,
        replace: str,
        sheet_name: Optional[str] = None,
        column: Optional[str] = None,
        case_sensitive: bool = False,
        whole_cell: bool = False,
        use_regex: bool = False,
        recursive: bool = False,
        max_files: int = 50,
    ) -> str:
        """Run find-and-replace across every .xlsx file in a folder.

        For each .xlsx: download -> find_and_replace -> upload.
        Native Google Sheets in the folder are skipped (use Sheets tools).
        Files where zero replacements happen are still re-uploaded UNLESS
        skipped on count == 0 (we skip upload to save bandwidth).

        Args:
            folder_id: Drive folder ID.
            find: Text/pattern to search.
            replace: Replacement.
            sheet_name: Restrict to a specific sheet name (applied per file).
                None means every sheet of every file.
            column: Restrict to a column letter or 1-based index string.
            case_sensitive: Case sensitivity flag.
            whole_cell: Require exact cell match.
            use_regex: Treat `find` as regex.
            recursive: Walk subfolders too.
            max_files: Hard cap on files processed.

        Returns:
            JSON: per-file replacement counts and totals.
        """
        try:
            listing = json.loads(
                self.list_excels_in_folder(
                    folder_id,
                    include_sheets=False,
                    recursive=recursive,
                    max_results=max_files,
                )
            )
            files = listing.get("files", [])
            if not files:
                return json.dumps(
                    {"processed": 0, "note": "No .xlsx files found in the folder."},
                    ensure_ascii=False,
                )

            per_file: list[dict[str, Any]] = []
            total_replacements = 0
            uploaded = 0
            skipped = 0
            errors: list[dict[str, str]] = []

            for f in files:
                fid = f["id"]
                fname = f["name"]
                try:
                    dl = self.download_excel(fid)
                    if dl.startswith(("Download error", "File is not")):
                        errors.append({"file": fname, "error": dl})
                        continue

                    rep = self.find_and_replace_excel(
                        file_id=fid,
                        find=find,
                        replace=replace,
                        sheet_name=sheet_name,
                        column=column,
                        case_sensitive=case_sensitive,
                        whole_cell=whole_cell,
                        use_regex=use_regex,
                    )
                    try:
                        rep_data = json.loads(rep)
                    except json.JSONDecodeError:
                        errors.append({"file": fname, "error": rep})
                        continue

                    count = rep_data.get("total", 0)
                    total_replacements += count

                    if count > 0:
                        up = self.upload_excel(fid)
                        if up.startswith("Upload error"):
                            errors.append({"file": fname, "error": up})
                            continue
                        uploaded += 1
                    else:
                        skipped += 1

                    per_file.append(
                        {
                            "file": fname,
                            "id": fid,
                            "replacements": count,
                            "uploaded": count > 0,
                        }
                    )
                except Exception as inner:
                    errors.append({"file": fname, "error": str(inner)})

            return json.dumps(
                {
                    "processed": len(per_file),
                    "uploaded": uploaded,
                    "skipped_no_match": skipped,
                    "total_replacements": total_replacements,
                    "per_file": per_file,
                    "errors": errors,
                },
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Bulk folder operation error: {e}"

    # ---------- Drive file management ----------

    def move_drive_file_to_trash(self, file_id: str) -> str:
        """Move a Drive file (Sheets or .xlsx) to Drive Trash.

        Trashed files are recoverable from Drive Trash for ~30 days; this
        is the safe way to remove a file. We deliberately do NOT expose a
        permanent-delete tool — irreversible deletion should be done by
        the user through the Drive UI.

        Args:
            file_id: Drive file ID.

        Returns:
            JSON: {file_id, name, trashed: true} or an error message.
        """
        try:
            updated = (
                self.drive.files()
                .update(
                    fileId=file_id,
                    body={"trashed": True},
                    fields="id, name, trashed",
                )
                .execute()
            )
            return json.dumps(
                {
                    "file_id": updated.get("id"),
                    "name": updated.get("name"),
                    "trashed": updated.get("trashed", False),
                    "note": "Moved to Drive Trash. Recoverable for ~30 days.",
                },
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Trash error: {e}"

    def rename_drive_file(self, file_id: str, new_name: str) -> str:
        """Rename any file on Drive (Sheets or .xlsx).

        Args:
            file_id: Drive file ID.
            new_name: New display name.

        Returns:
            JSON: {file_id, name} or error.
        """
        try:
            updated = (
                self.drive.files()
                .update(
                    fileId=file_id,
                    body={"name": new_name},
                    fields="id, name",
                )
                .execute()
            )
            return json.dumps(
                {"file_id": updated.get("id"), "name": updated.get("name")},
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Rename error: {e}"

    def copy_drive_file(
        self,
        file_id: str,
        new_name: Optional[str] = None,
        target_folder_id: Optional[str] = None,
    ) -> str:
        """Duplicate any Drive file (Sheets or .xlsx).

        Args:
            file_id: Source file ID.
            new_name: Name for the copy. None -> "Copy of <original>".
            target_folder_id: Drive folder to put the copy in. None keeps the
                copy in the source's parent folder.

        Returns:
            JSON: {file_id, name, mimeType, url} or error.
        """
        try:
            body: dict[str, Any] = {}
            if new_name:
                body["name"] = new_name
            if target_folder_id:
                body["parents"] = [target_folder_id]
            new_file = (
                self.drive.files()
                .copy(
                    fileId=file_id,
                    body=body or None,
                    fields="id, name, mimeType, webViewLink",
                )
                .execute()
            )
            return json.dumps(
                {
                    "file_id": new_file.get("id"),
                    "name": new_file.get("name"),
                    "mimeType": new_file.get("mimeType"),
                    "url": new_file.get("webViewLink"),
                },
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Copy error: {e}"

    # ---------- Sheet tab management (Drive .xlsx) ----------

    def rename_excel_sheet_tab(
        self, file_id: str, old_name: str, new_name: str
    ) -> str:
        """Rename a sheet (tab) inside a downloaded .xlsx file.

        Requires `download_excel(file_id)` first. After this, call
        `upload_excel(file_id)` to push changes back to Drive.

        Args:
            file_id: Drive file ID.
            old_name: Current sheet name.
            new_name: New sheet name.

        Returns:
            Success message or error.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path)
            if old_name not in wb.sheetnames:
                wb.close()
                return f"Sheet not found: {old_name}. Available: {wb.sheetnames}"
            if new_name in wb.sheetnames and new_name != old_name:
                wb.close()
                return f"A sheet named {new_name!r} already exists."
            wb[old_name].title = new_name
            wb.save(path)
            wb.close()
            return f"Sheet renamed {old_name!r} -> {new_name!r}. Don't forget upload_excel."
        except Exception as e:
            return f"Rename sheet error: {e}"

    def delete_excel_sheet_tab(self, file_id: str, sheet_name: str) -> str:
        """Delete a sheet (tab) from a downloaded .xlsx file.

        Requires `download_excel(file_id)` first. Refuses to delete the last
        remaining sheet.

        Args:
            file_id: Drive file ID.
            sheet_name: Sheet to remove.

        Returns:
            Success message or error.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            if len(wb.sheetnames) == 1:
                wb.close()
                return (
                    f"Cannot delete {sheet_name!r}: it is the only sheet in "
                    "the workbook. Create another sheet first."
                )
            del wb[sheet_name]
            wb.save(path)
            wb.close()
            return f"Sheet {sheet_name!r} deleted. Don't forget upload_excel."
        except Exception as e:
            return f"Delete sheet error: {e}"

    # ---------- Column operations + sort (Drive .xlsx) ----------

    def sort_excel_by_column(
        self,
        file_id: str,
        sheet_name: str,
        column: str,
        ascending: bool = True,
        has_header: bool = True,
    ) -> str:
        """Sort rows by column in a downloaded .xlsx file.

        Requires `download_excel(file_id)` first; remember to call
        `upload_excel(file_id)` afterwards.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
            col_idx = _resolve_column(column)

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                wb.close()
                return f"{sheet_name}: empty sheet, nothing to sort."
            header = rows[0] if has_header else None
            body = rows[1:] if has_header else rows

            def key(row: tuple) -> Any:
                v = row[col_idx - 1] if col_idx - 1 < len(row) else None
                if v is None:
                    return (1, "")
                if isinstance(v, (int, float)):
                    return (0, v)
                return (0, str(v).lower())

            body.sort(key=key, reverse=not ascending)
            ws.delete_rows(1, ws.max_row)
            if header is not None:
                ws.append(list(header))
            for r in body:
                ws.append(list(r))
            wb.save(path)
            wb.close()
            return (
                f"{sheet_name}: sorted by column {column} "
                f"({'asc' if ascending else 'desc'}), {len(body)} rows. "
                "Don't forget upload_excel."
            )
        except Exception as e:
            return f"Sort error: {e}"

    def delete_excel_columns(
        self, file_id: str, sheet_name: str, start_col: str, count: int = 1
    ) -> str:
        """Delete one or more columns from a downloaded .xlsx file."""
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            col_idx = _resolve_column(start_col)
            wb[sheet_name].delete_cols(col_idx, count)
            wb.save(path)
            wb.close()
            return f"{sheet_name}: {count} column(s) deleted from {start_col}. Don't forget upload_excel."
        except Exception as e:
            return f"Delete columns error: {e}"

    def insert_excel_columns(
        self, file_id: str, sheet_name: str, before_col: str, count: int = 1
    ) -> str:
        """Insert blank columns into a downloaded .xlsx file."""
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            col_idx = _resolve_column(before_col)
            wb[sheet_name].insert_cols(col_idx, count)
            wb.save(path)
            wb.close()
            return f"{sheet_name}: {count} blank column(s) inserted before {before_col}. Don't forget upload_excel."
        except Exception as e:
            return f"Insert columns error: {e}"

    # ---------- Analytical + export + creation (Drive .xlsx) ----------

    def describe_excel(
        self, file_id: str, sheet_name: Optional[str] = None, sample_size: int = 5
    ) -> str:
        """Produce a per-column overview of a downloaded .xlsx file.

        Requires `download_excel(file_id)` first. Reports header, inferred
        type, total/non-empty/unique counts, min/max for numeric columns,
        and a few sample values for each column of each sheet.

        Args:
            file_id: Drive file ID.
            sheet_name: Restrict to one sheet. None -> describe all sheets.
            sample_size: Number of sample values per column.

        Returns:
            JSON summary.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            sheets = [sheet_name] if sheet_name else wb.sheetnames
            for s in sheets:
                if s not in wb.sheetnames:
                    wb.close()
                    return f"Sheet not found: {s}. Available: {wb.sheetnames}"
            result: dict[str, Any] = {"file_id": file_id, "sheets": {}}
            for s in sheets:
                ws = wb[s]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    result["sheets"][s] = {"rows": 0, "columns": []}
                    continue
                header = rows[0]
                body = rows[1:]
                col_count = max(len(r) for r in rows) if rows else 0

                columns_info: list[dict[str, Any]] = []
                for col_idx in range(col_count):
                    values = [r[col_idx] if col_idx < len(r) else None for r in body]
                    non_empty = [
                        v for v in values
                        if v is not None and not (isinstance(v, str) and not v.strip())
                    ]
                    nums: list[float] = []
                    non_numeric = 0
                    for v in non_empty:
                        try:
                            nums.append(float(v))
                        except (TypeError, ValueError):
                            non_numeric += 1
                    inferred = (
                        "numeric"
                        if nums and non_numeric == 0
                        else "mixed" if nums and non_numeric > 0
                        else "text" if non_empty
                        else "empty"
                    )
                    info: dict[str, Any] = {
                        "header": str(header[col_idx]) if col_idx < len(header) else None,
                        "letter": get_column_letter(col_idx + 1),
                        "type": inferred,
                        "total": len(values),
                        "non_empty": len(non_empty),
                        "empty": len(values) - len(non_empty),
                        "unique": len({str(v) for v in non_empty}),
                        "samples": [str(v) for v in non_empty[:sample_size]],
                    }
                    if nums:
                        info["min"] = min(nums)
                        info["max"] = max(nums)
                        info["mean"] = sum(nums) / len(nums)
                    columns_info.append(info)
                result["sheets"][s] = {"rows": len(body), "columns": columns_info}
            wb.close()
            return json.dumps(result, ensure_ascii=False, default=str)
        except Exception as e:
            return f"Describe error: {e}"

    def export_drive_excel_to_csv(
        self, file_id: str, sheet_name: str, output_path: str
    ) -> str:
        """Export a sheet of a downloaded Drive .xlsx as CSV on the local disk.

        Requires `download_excel(file_id)` first. The CSV is written to the
        user-provided absolute (or cwd-relative) path. Does NOT upload back
        to Drive.

        Args:
            file_id: Drive file ID.
            sheet_name: Sheet to export.
            output_path: Local destination .csv path.

        Returns:
            JSON or error.
        """
        path = _local_path(file_id)
        if not path.exists():
            return f"Local file not found: {path}. Call download_excel first."
        try:
            import csv
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
            out = Path(output_path).expanduser().resolve()
            if not str(out).lower().endswith(".csv"):
                out = out.with_suffix(".csv")
            out.parent.mkdir(parents=True, exist_ok=True)
            n = 0
            with out.open("w", encoding="utf-8", newline="") as fh:
                writer = csv.writer(fh)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(["" if v is None else v for v in row])
                    n += 1
            wb.close()
            return json.dumps(
                {"sheet": sheet_name, "output": str(out), "rows": n},
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Export error: {e}"

    def create_drive_xlsx_file(
        self,
        name: str,
        sheet_name: str = "Sheet1",
        target_folder_id: Optional[str] = None,
    ) -> str:
        """Create a brand-new empty .xlsx file on Drive (not Google Sheets).

        Creates the workbook locally, uploads it as a new .xlsx file, then
        downloads it back to `excel_workdir/` so the agent can edit it with
        the existing tools without an extra round-trip.

        Args:
            name: Desired Drive display name (.xlsx is appended if missing).
            sheet_name: Name of the initial sheet/tab.
            target_folder_id: Drive folder to put it in. None -> "My Drive".

        Returns:
            JSON: {file_id, name, url} or error.
        """
        try:
            import io
            from openpyxl import Workbook
            from googleapiclient.http import MediaIoBaseUpload

            display_name = name if name.lower().endswith(".xlsx") else f"{name}.xlsx"

            # Build the workbook in memory
            wb = Workbook()
            wb.active.title = sheet_name
            buf = io.BytesIO()
            wb.save(buf)
            wb.close()
            buf.seek(0)

            meta: dict[str, Any] = {"name": display_name, "mimeType": EXCEL_MIME}
            if target_folder_id:
                meta["parents"] = [target_folder_id]
            media = MediaIoBaseUpload(buf, mimetype=EXCEL_MIME, resumable=False)
            new_file = (
                self.drive.files()
                .create(body=meta, media_body=media, fields="id, name, webViewLink")
                .execute()
            )
            new_id = new_file["id"]

            # Cache locally so the agent can edit immediately
            buf.seek(0)
            _local_path(new_id).write_bytes(buf.getvalue())

            return json.dumps(
                {
                    "file_id": new_id,
                    "name": new_file.get("name"),
                    "url": new_file.get("webViewLink"),
                    "note": (
                        "New .xlsx created on Drive AND cached locally; you can "
                        "use update_excel_cell / update_excel_range straight away."
                    ),
                },
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Create error: {e}"


# ---------- Module-level helpers ----------


def _resolve_column(column: str) -> int:
    """Convert a column letter ("B") or 1-based numeric string ("2") to an index."""
    s = str(column).strip()
    if s.isdigit():
        return int(s)
    return column_index_from_string(s.upper())


def _to_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _match_condition(
    cell_value: Any, operator: str, value: Optional[Union[str, float, int]]
) -> bool:
    """Evaluate `cell_value <operator> value`. Used by filter/delete tools."""
    op = operator.strip().lower()

    if op in {"empty", "is_empty"}:
        return cell_value is None or (isinstance(cell_value, str) and not cell_value.strip())
    if op in {"not_empty", "is_not_empty"}:
        return not (cell_value is None or (isinstance(cell_value, str) and not cell_value.strip()))

    if op in {"contains", "not_contains"}:
        if cell_value is None or value is None:
            return op == "not_contains"
        hit = str(value).lower() in str(cell_value).lower()
        return hit if op == "contains" else not hit

    if op in {"==", "=", "equals"}:
        if cell_value is None and value is None:
            return True
        if cell_value is None or value is None:
            return False
        a, b = _to_number(cell_value), _to_number(value)
        if a is not None and b is not None:
            return a == b
        return str(cell_value) == str(value)
    if op in {"!=", "<>", "not_equals"}:
        return not _match_condition(cell_value, "==", value)

    a = _to_number(cell_value)
    b = _to_number(value)
    if a is None or b is None:
        return False
    if op in {">", "gt"}:
        return a > b
    if op in {"<", "lt"}:
        return a < b
    if op in {">=", "gte"}:
        return a >= b
    if op in {"<=", "lte"}:
        return a <= b
    return False
