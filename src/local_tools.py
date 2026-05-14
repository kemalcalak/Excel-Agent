"""Local-disk .xlsx editing — opt-in folder workspace.

By default the agent has NO access to the local filesystem. The user must
explicitly call `open_local_folder(path)` to activate a workspace. Once a
folder is open every local_* tool resolves filenames against it and refuses
to escape (prevents path traversal).

This module duplicates a few openpyxl operations that also exist in
`excel_tools.py`; we keep them separate because the Drive-based tools and
local tools have different semantics (Drive ID vs filename) and the agent
needs distinct tool names to choose the right path.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, List, Optional, Union

from agno.tools import Toolkit
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.excel_tools import _match_condition, _resolve_column

# Directories skipped by recursive walks and folder-wide bulk operations.
# These are dev/build artefacts that almost never hold user spreadsheets;
# walking them just blows past max_results and slows things down. Extend if
# the user has another well-known noise folder.
IGNORE_DIRS: frozenset[str] = frozenset(
    {
        ".venv",
        "venv",
        "env",
        "__pycache__",
        ".git",
        ".hg",
        ".svn",
        "node_modules",
        ".idea",
        ".vscode",
        "dist",
        "build",
        ".pytest_cache",
        ".mypy_cache",
        ".ruff_cache",
        ".tox",
        ".cache",
        "excel_workdir",  # our own Drive download cache
    }
)


def _is_ignored(rel_path: Path) -> bool:
    """True if any path component matches an entry in IGNORE_DIRS."""
    return any(part in IGNORE_DIRS for part in rel_path.parts)


class LocalExcelTools(Toolkit):
    """Local filesystem .xlsx workspace.

    State:
        active_folder: Path | None. Set via open_local_folder(); cleared via
            close_local_folder(). Every other tool refuses to run when None.
    """

    def __init__(self) -> None:
        self.active_folder: Optional[Path] = None
        super().__init__(
            name="local_excel_tools",
            tools=[
                # Workspace lifecycle
                self.open_local_folder,
                self.close_local_folder,
                self.show_local_folder,
                # Filesystem inspection
                self.list_local_folder,
                # Excel inspection
                self.list_local_sheet_names,
                self.read_local_excel,
                self.find_cells_local_excel,
                self.filter_local_excel_rows,
                self.column_summary_local_excel,
                # Excel editing — single
                self.create_local_excel_file,
                self.update_local_excel_cell,
                self.update_local_excel_range,
                self.append_local_excel_row,
                self.create_local_excel_sheet_tab,
                self.set_formula_local_excel,
                # Excel editing — bulk
                self.find_and_replace_local_excel,
                self.delete_local_excel_rows,
                self.delete_local_excel_rows_where,
                self.insert_local_excel_rows,
                # Folder-wide bulk
                self.bulk_find_replace_in_local_folder,
                # File management (Recycle Bin)
                self.move_local_excel_file_to_trash,
                # Sheet tab management
                self.rename_local_excel_sheet_tab,
                self.delete_local_excel_sheet_tab,
                # File management
                self.rename_local_excel_file,
                self.copy_local_excel_file,
                self.move_local_excel_file,
                # Column operations + sort
                self.sort_local_excel_by_column,
                self.delete_local_excel_columns,
                self.insert_local_excel_columns,
                # Analytical + export
                self.describe_local_excel,
                self.search_in_all_local_files,
                self.export_local_excel_to_csv,
            ],
        )

    # ---------- Workspace lifecycle ----------

    # Default workspace shipped with the project. Users drop .xlsx files here
    # so the agent has a tidy, well-known location to look. Any other path can
    # still be passed explicitly.
    DEFAULT_FOLDER = "workbooks"

    def open_local_folder(self, path: str = "") -> str:
        """Activate a folder on local disk as the workspace.

        The agent has NO access to any local file until this is called.
        Subsequent local_* tools resolve filenames relative to this folder.

        Args:
            path: Absolute or relative directory path. Relative paths resolve
                against the current working directory. When empty, the
                project's default `workbooks/` folder is opened.

        Returns:
            JSON: {active_folder, contents} on success; error string otherwise.
        """
        target = path or self.DEFAULT_FOLDER
        try:
            p = Path(target).expanduser().resolve()
        except Exception as e:
            return f"Path resolution error: {e}"
        if not p.exists():
            return f"Path does not exist: {p}"
        if not p.is_dir():
            return f"Not a directory: {p}"

        self.active_folder = p
        is_default = path == "" or path == self.DEFAULT_FOLDER
        return json.dumps(
            {
                "active_folder": str(p),
                "is_default_folder": is_default,
                "contents": self._ls(p),
                "note": "Folder opened. Use local_* tools to read/edit files inside it.",
            },
            ensure_ascii=False,
        )

    def close_local_folder(self) -> str:
        """Deactivate the local workspace. Subsequent local_* tools will refuse.

        Returns:
            Confirmation message.
        """
        prev = self.active_folder
        self.active_folder = None
        if prev is None:
            return "No folder was open."
        return f"Closed local folder: {prev}"

    def show_local_folder(self) -> str:
        """Return the currently active folder path, or a hint when none is open."""
        if self.active_folder is None:
            return "No folder is open. Call open_local_folder(path) to activate one."
        return str(self.active_folder)

    # ---------- Filesystem inspection ----------

    def list_local_folder(
        self,
        subpath: str = "",
        recursive: bool = False,
        max_results: int = 500,
    ) -> str:
        """List entries inside the active workspace (or a subpath of it).

        Args:
            subpath: Optional path relative to active_folder. Empty -> root.
            recursive: If True, walk subdirectories too. Entry names are
                returned as relative paths (forward-slash separated).
            max_results: Cap on entries when recursive.

        Returns:
            JSON: {path, recursive, entries: [{name, type, size, ext}]}.
        """
        if self.active_folder is None:
            return "No folder is open. Call open_local_folder(path) first."
        try:
            target = self._resolve_subpath(subpath) if subpath else self.active_folder
        except RuntimeError as e:
            return str(e)
        if not target.is_dir():
            return f"Not a directory: {target}"

        entries = self._walk(target, max_results) if recursive else self._ls(target)
        return json.dumps(
            {
                "path": str(target),
                "recursive": recursive,
                "entries": entries,
                "truncated": recursive and len(entries) >= max_results,
            },
            ensure_ascii=False,
        )

    # ---------- Excel inspection ----------

    def list_local_sheet_names(self, filename: str) -> str:
        """List sheet (tab) names inside a local .xlsx file.

        Args:
            filename: Path relative to active_folder, e.g. "report.xlsx" or
                "sub/report.xlsx".

        Returns:
            JSON array of sheet names.
        """
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path, read_only=True, data_only=False)
            names = wb.sheetnames
            wb.close()
            return json.dumps(names, ensure_ascii=False)
        except Exception as e:
            return f"Read error: {e}"

    def read_local_excel(
        self,
        filename: str,
        sheet_name: str,
        cell_range: Optional[str] = None,
        values_only: bool = True,
    ) -> str:
        """Read cells from a local .xlsx file.

        Args:
            filename: Path relative to active_folder.
            sheet_name: Sheet/tab name.
            cell_range: Range like "A1:D20"; None returns the whole used area.
            values_only: True -> only values; False -> textual cell repr.

        Returns:
            JSON 2D list.
        """
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
            rows: List[List[Any]] = []
            if cell_range:
                target = ws[cell_range]
                if not isinstance(target, tuple):
                    target = (target,)
                for row in target:
                    if not isinstance(row, tuple):
                        row = (row,)
                    rows.append([c.value for c in row])
            else:
                for row in ws.iter_rows(values_only=values_only):
                    rows.append(list(row) if isinstance(row, tuple) else [row])
            wb.close()
            return json.dumps(rows, ensure_ascii=False, default=str)
        except Exception as e:
            return f"Read error: {e}"

    def find_cells_local_excel(
        self,
        filename: str,
        sheet_name: str,
        pattern: str,
        case_sensitive: bool = False,
        whole_cell: bool = False,
        use_regex: bool = False,
        max_results: int = 200,
    ) -> str:
        """Locate cells whose value matches a pattern in a local .xlsx file.

        Args:
            filename: Path relative to active_folder.
            sheet_name: Sheet to scan.
            pattern: Substring or regex.
            case_sensitive: Case sensitivity flag.
            whole_cell: Require exact cell match.
            use_regex: Interpret pattern as regex.
            max_results: Cap on matches.

        Returns:
            JSON list of {cell, value}.
        """
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]

            regex: Optional[re.Pattern[str]] = None
            if use_regex:
                flags = 0 if case_sensitive else re.IGNORECASE
                regex = re.compile(pattern, flags)
            needle = pattern if case_sensitive else pattern.lower()

            hits: list[dict[str, Any]] = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value)
                    cmp_text = text if case_sensitive else text.lower()
                    match = False
                    if use_regex and regex:
                        match = bool(regex.fullmatch(text) if whole_cell else regex.search(text))
                    else:
                        match = (cmp_text == needle) if whole_cell else (needle in cmp_text)
                    if match:
                        hits.append({"cell": cell.coordinate, "value": text})
                        if len(hits) >= max_results:
                            break
                if len(hits) >= max_results:
                    break
            wb.close()
            return json.dumps(
                {"matches": hits, "count": len(hits), "truncated": len(hits) >= max_results},
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Search error: {e}"

    def filter_local_excel_rows(
        self,
        filename: str,
        sheet_name: str,
        column: str,
        operator: str,
        value: Optional[Union[str, float, int]] = None,
        has_header: bool = True,
        max_results: int = 200,
    ) -> str:
        """Return local-file rows matching a column condition.

        Args / operators identical to filter_rows_excel in ExcelTools.
        """
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
            col_idx = _resolve_column(column)

            header: Optional[List[Any]] = None
            rows: list[list[Any]] = []
            start = 2 if has_header else 1
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if has_header and i == 1:
                    header = list(row)
                    continue
                if i < start:
                    continue
                cell_value = row[col_idx - 1] if col_idx - 1 < len(row) else None
                if _match_condition(cell_value, operator, value):
                    rows.append(list(row))
                    if len(rows) >= max_results:
                        break
            wb.close()
            return json.dumps(
                {
                    "header": header,
                    "rows": rows,
                    "count": len(rows),
                    "truncated": len(rows) >= max_results,
                },
                ensure_ascii=False,
                default=str,
            )
        except Exception as e:
            return f"Filter error: {e}"

    def column_summary_local_excel(
        self,
        filename: str,
        sheet_name: str,
        column: str,
        has_header: bool = True,
    ) -> str:
        """Compute count/sum/mean/min/max for a column in a local .xlsx file."""
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
            col_idx = _resolve_column(column)

            nums: list[float] = []
            empty = 0
            non_numeric = 0
            total = 0
            for i, row in enumerate(
                ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True), start=1
            ):
                if has_header and i == 1:
                    continue
                total += 1
                v = row[0]
                if v is None or (isinstance(v, str) and not v.strip()):
                    empty += 1
                    continue
                try:
                    nums.append(float(v))
                except (TypeError, ValueError):
                    non_numeric += 1
            wb.close()

            summary: dict[str, Any] = {
                "count": total,
                "numeric_count": len(nums),
                "non_numeric_count": non_numeric,
                "empty_count": empty,
            }
            if nums:
                summary.update(
                    sum=sum(nums),
                    mean=sum(nums) / len(nums),
                    min=min(nums),
                    max=max(nums),
                )
            return json.dumps(summary, ensure_ascii=False)
        except Exception as e:
            return f"Summary error: {e}"

    # ---------- Excel editing ----------

    def create_local_excel_file(
        self, filename: str, sheet_name: str = "Sheet1", overwrite: bool = False
    ) -> str:
        """Create a brand-new empty .xlsx file inside the active workspace.

        Required precondition for the other local editing tools: they all
        operate on a file that already exists. Use this when the user asks
        for a new file.

        Args:
            filename: Path relative to active_folder. ".xlsx" is appended if
                missing. Cannot escape the workspace.
            sheet_name: Name of the first sheet/tab. Defaults to "Sheet1".
            overwrite: If False (default) and the file already exists, the
                call fails. Set True to replace the file (destructive).

        Returns:
            JSON: {file, sheet, full_path} on success; error string otherwise.
        """
        if self.active_folder is None:
            return "No folder is open. Call open_local_folder(path) first."
        try:
            from openpyxl import Workbook

            target_name = filename if filename.lower().endswith(".xlsx") else f"{filename}.xlsx"
            candidate = (self.active_folder / target_name).resolve()
            try:
                candidate.relative_to(self.active_folder)
            except ValueError:
                return f"Path escapes the active folder: {filename}"

            if candidate.exists() and not overwrite:
                return (
                    f"File already exists: {candidate.name}. Pass overwrite=True "
                    "to replace it, or pick a different filename."
                )

            candidate.parent.mkdir(parents=True, exist_ok=True)
            wb = Workbook()
            default_ws = wb.active
            default_ws.title = sheet_name
            wb.save(candidate)
            wb.close()
            return json.dumps(
                {
                    "file": target_name,
                    "sheet": sheet_name,
                    "full_path": str(candidate),
                    "note": "Empty .xlsx created. Use update_local_excel_range / append_local_excel_row to add data.",
                },
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Create error: {e}"

    def update_local_excel_cell(
        self, filename: str, sheet_name: str, cell: str, value: Any
    ) -> str:
        """Set one cell in a local .xlsx file. Saves in place."""
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            wb[sheet_name][cell] = value
            wb.save(path)
            wb.close()
            return f"{filename}!{sheet_name}!{cell} = {value!r} (saved in place)"
        except Exception as e:
            return f"Write error: {e}"

    def update_local_excel_range(
        self,
        filename: str,
        sheet_name: str,
        start_cell: str,
        data: List[List[Any]],
    ) -> str:
        """Write a 2D block into a local .xlsx file starting at start_cell."""
        try:
            from openpyxl.utils.cell import range_boundaries
            path = self._resolve_file(filename)
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
            start_col, start_row, _, _ = range_boundaries(start_cell + ":" + start_cell)
            for r, row in enumerate(data):
                for c, val in enumerate(row):
                    ws.cell(row=start_row + r, column=start_col + c, value=val)
            wb.save(path)
            wb.close()
            end_cell = (
                f"{get_column_letter(start_col + max((len(r) for r in data), default=1) - 1)}"
                f"{start_row + len(data) - 1}"
            )
            return f"{filename}!{sheet_name}!{start_cell}:{end_cell} written (saved in place)"
        except Exception as e:
            return f"Write error: {e}"

    def append_local_excel_row(
        self, filename: str, sheet_name: str, row_values: List[Any]
    ) -> str:
        """Append a row at the bottom of a sheet in a local .xlsx file."""
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            wb[sheet_name].append(row_values)
            wb.save(path)
            wb.close()
            return f"Row appended to {filename}!{sheet_name} (saved in place)"
        except Exception as e:
            return f"Write error: {e}"

    def create_local_excel_sheet_tab(self, filename: str, sheet_name: str) -> str:
        """Add a new sheet tab to a local .xlsx file."""
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path)
            if sheet_name in wb.sheetnames:
                return f"Sheet already exists: {sheet_name}"
            wb.create_sheet(title=sheet_name)
            wb.save(path)
            wb.close()
            return f"New sheet added: {filename}!{sheet_name} (saved in place)"
        except Exception as e:
            return f"Sheet creation error: {e}"

    def set_formula_local_excel(
        self,
        filename: str,
        sheet_name: str,
        target_column: str,
        formula_template: str,
        start_row: int,
        end_row: int,
    ) -> str:
        """Apply a templated formula to a column range, substituting `{row}`."""
        if start_row > end_row:
            return "start_row cannot be greater than end_row."
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
            col_idx = _resolve_column(target_column)
            for r in range(start_row, end_row + 1):
                ws.cell(
                    row=r, column=col_idx, value=formula_template.replace("{row}", str(r))
                )
            wb.save(path)
            wb.close()
            col_letter = get_column_letter(col_idx)
            return (
                f"Formula written to {filename}!{sheet_name}!"
                f"{col_letter}{start_row}:{col_letter}{end_row} (saved in place)"
            )
        except Exception as e:
            return f"Formula write error: {e}"

    def find_and_replace_local_excel(
        self,
        filename: str,
        find: str,
        replace: str,
        sheet_name: Optional[str] = None,
        column: Optional[str] = None,
        case_sensitive: bool = False,
        whole_cell: bool = False,
        use_regex: bool = False,
    ) -> str:
        """Bulk find-and-replace inside a local .xlsx file. Saves in place."""
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path)
            sheets = [sheet_name] if sheet_name else wb.sheetnames
            for s in sheets:
                if s not in wb.sheetnames:
                    return f"Sheet not found: {s}. Available: {wb.sheetnames}"

            col_idx = _resolve_column(column) if column else None
            pattern: Optional[re.Pattern[str]] = None
            if use_regex:
                flags = 0 if case_sensitive else re.IGNORECASE
                pattern = re.compile(find, flags)

            counts: dict[str, int] = {}
            for s in sheets:
                ws = wb[s]
                n = 0
                for row in ws.iter_rows():
                    for cell in row:
                        if col_idx is not None and cell.column != col_idx:
                            continue
                        if cell.value is None:
                            continue
                        text = str(cell.value)
                        new_text: Optional[str] = None
                        if whole_cell:
                            if use_regex and pattern and pattern.fullmatch(text):
                                new_text = replace
                            elif not use_regex and (
                                (case_sensitive and text == find)
                                or (not case_sensitive and text.lower() == find.lower())
                            ):
                                new_text = replace
                        else:
                            if use_regex and pattern:
                                replaced = pattern.sub(replace, text)
                                if replaced != text:
                                    new_text = replaced
                            elif not use_regex:
                                if case_sensitive and find in text:
                                    new_text = text.replace(find, replace)
                                elif not case_sensitive and find.lower() in text.lower():
                                    new_text = re.sub(
                                        re.escape(find), replace, text, flags=re.IGNORECASE
                                    )
                        if new_text is not None:
                            cell.value = new_text
                            n += 1
                counts[s] = n

            wb.save(path)
            wb.close()
            total = sum(counts.values())
            return json.dumps(
                {
                    "file": filename,
                    "replacements_per_sheet": counts,
                    "total": total,
                    "note": "Saved in place.",
                },
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Find/replace error: {e}"

    def delete_local_excel_rows(
        self, filename: str, sheet_name: str, start_row: int, count: int = 1
    ) -> str:
        """Delete a contiguous row range in a local .xlsx file."""
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            wb[sheet_name].delete_rows(start_row, count)
            wb.save(path)
            wb.close()
            return f"{filename}!{sheet_name}: {count} rows deleted from row {start_row} (saved in place)"
        except Exception as e:
            return f"Delete error: {e}"

    def delete_local_excel_rows_where(
        self,
        filename: str,
        sheet_name: str,
        column: str,
        operator: str,
        value: Optional[Union[str, float, int]] = None,
        has_header: bool = True,
    ) -> str:
        """Delete rows where a column value matches a condition."""
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
            col_idx = _resolve_column(column)
            rows_to_delete: list[int] = []
            start = 2 if has_header else 1
            for row_idx in range(start, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if _match_condition(cell_value, operator, value):
                    rows_to_delete.append(row_idx)
            for r in reversed(rows_to_delete):
                ws.delete_rows(r, 1)
            wb.save(path)
            wb.close()
            return json.dumps(
                {
                    "file": filename,
                    "deleted": len(rows_to_delete),
                    "remaining": ws.max_row,
                    "note": "Saved in place.",
                },
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Conditional delete error: {e}"

    def insert_local_excel_rows(
        self, filename: str, sheet_name: str, before_row: int, count: int = 1
    ) -> str:
        """Insert blank rows above the given 1-based row index."""
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            wb[sheet_name].insert_rows(before_row, count)
            wb.save(path)
            wb.close()
            return f"{filename}!{sheet_name}: {count} blank rows inserted before row {before_row} (saved in place)"
        except Exception as e:
            return f"Row insert error: {e}"

    # ---------- Folder-wide bulk operations ----------

    def bulk_find_replace_in_local_folder(
        self,
        find: str,
        replace: str,
        sheet_name: Optional[str] = None,
        column: Optional[str] = None,
        case_sensitive: bool = False,
        whole_cell: bool = False,
        use_regex: bool = False,
        recursive: bool = False,
        max_files: int = 100,
    ) -> str:
        """Run find-and-replace across every .xlsx in the active folder.

        Mirrors `bulk_find_replace_in_folder` (Drive) but on local disk.
        Each file is saved in place.

        Args:
            find: Text or regex pattern to search for.
            replace: Replacement.
            sheet_name: Restrict to a specific sheet name (applied per file).
                None means every sheet of every file.
            column: Restrict to a column letter or 1-based index string.
            case_sensitive: Case sensitivity flag.
            whole_cell: Require exact cell match.
            use_regex: Treat `find` as a regex.
            recursive: Walk subdirectories too.
            max_files: Hard cap on files processed.

        Returns:
            JSON: per-file replacement counts and totals.
        """
        if self.active_folder is None:
            return "No folder is open. Call open_local_folder(path) first."
        try:
            pattern = "**/*.xlsx" if recursive else "*.xlsx"
            files = [
                p
                for p in self.active_folder.glob(pattern)
                if p.is_file()
                and not p.name.startswith("~$")  # skip Excel lock files
                and not _is_ignored(p.relative_to(self.active_folder))
            ][:max_files]

            if not files:
                return json.dumps(
                    {"processed": 0, "note": "No .xlsx files found in the active folder."},
                    ensure_ascii=False,
                )

            per_file: list[dict[str, Any]] = []
            total_replacements = 0
            modified = 0
            errors: list[dict[str, str]] = []

            for fpath in files:
                rel = fpath.relative_to(self.active_folder).as_posix()
                try:
                    rep = self.find_and_replace_local_excel(
                        filename=rel,
                        find=find,
                        replace=replace,
                        sheet_name=sheet_name,
                        column=column,
                        case_sensitive=case_sensitive,
                        whole_cell=whole_cell,
                        use_regex=use_regex,
                    )
                    try:
                        rep_data = json.loads(rep)
                    except json.JSONDecodeError:
                        errors.append({"file": rel, "error": rep})
                        continue

                    count = rep_data.get("total", 0)
                    total_replacements += count
                    if count > 0:
                        modified += 1
                    per_file.append({"file": rel, "replacements": count})
                except Exception as inner:
                    errors.append({"file": rel, "error": str(inner)})

            return json.dumps(
                {
                    "processed": len(per_file),
                    "modified": modified,
                    "unchanged": len(per_file) - modified,
                    "total_replacements": total_replacements,
                    "per_file": per_file,
                    "errors": errors,
                },
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Bulk local operation error: {e}"

    # ---------- File management (Recycle Bin) ----------

    def move_local_excel_file_to_trash(self, filename: str) -> str:
        """Move a local .xlsx file to the OS Recycle Bin.

        Mirrors `move_drive_file_to_trash` on the Drive side: recoverable
        deletion via send2trash, which uses the Windows Recycle Bin (or
        the macOS/Linux equivalent). We deliberately do NOT expose a
        permanent-delete tool — irreversible removal must be done by the
        user through the OS file manager.

        Args:
            filename: Path relative to active_folder. Subpaths are allowed
                (e.g. "subdir/old.xlsx") but cannot escape the workspace.

        Returns:
            JSON: {file, full_path, trashed: true, note} on success;
            error string otherwise.
        """
        if self.active_folder is None:
            return "No folder is open. Call open_local_folder(path) first."
        try:
            from send2trash import send2trash

            try:
                path = self._resolve_file(filename)
            except RuntimeError as e:
                return str(e)

            send2trash(str(path))
            return json.dumps(
                {
                    "file": filename,
                    "full_path": str(path),
                    "trashed": True,
                    "note": "Moved to the OS Recycle Bin. Recoverable from there.",
                },
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Trash error: {e}"

    # ---------- Sheet tab management ----------

    def rename_local_excel_sheet_tab(
        self, filename: str, old_name: str, new_name: str
    ) -> str:
        """Rename a sheet (tab) inside a local .xlsx file.

        Args:
            filename: Path relative to active_folder.
            old_name: Current sheet name.
            new_name: New sheet name.

        Returns:
            Success message or error.
        """
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path)
            if old_name not in wb.sheetnames:
                return f"Sheet not found: {old_name}. Available: {wb.sheetnames}"
            if new_name in wb.sheetnames and new_name != old_name:
                return f"A sheet named {new_name!r} already exists."
            wb[old_name].title = new_name
            wb.save(path)
            wb.close()
            return f"{filename}: sheet renamed {old_name!r} -> {new_name!r} (saved in place)"
        except Exception as e:
            return f"Rename sheet error: {e}"

    def delete_local_excel_sheet_tab(self, filename: str, sheet_name: str) -> str:
        """Delete a sheet (tab) from a local .xlsx file.

        Refuses to delete the last remaining sheet because openpyxl requires
        at least one sheet in a workbook.

        Args:
            filename: Path relative to active_folder.
            sheet_name: Sheet to remove.

        Returns:
            Success message or error.
        """
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            if len(wb.sheetnames) == 1:
                return (
                    f"Cannot delete {sheet_name!r}: it is the only sheet in "
                    "the workbook. Create another sheet first."
                )
            del wb[sheet_name]
            wb.save(path)
            wb.close()
            return f"{filename}: sheet {sheet_name!r} deleted (saved in place)"
        except Exception as e:
            return f"Delete sheet error: {e}"

    # ---------- File management ----------

    def rename_local_excel_file(self, old_filename: str, new_filename: str) -> str:
        """Rename a .xlsx file inside the active workspace.

        Args:
            old_filename: Existing filename relative to active_folder.
            new_filename: New filename relative to active_folder. ".xlsx" is
                appended if missing.

        Returns:
            JSON with new path or error.
        """
        if self.active_folder is None:
            return "No folder is open. Call open_local_folder(path) first."
        try:
            src = self._resolve_file(old_filename)
            new_name = (
                new_filename if new_filename.lower().endswith(".xlsx") else f"{new_filename}.xlsx"
            )
            dst = (self.active_folder / new_name).resolve()
            try:
                dst.relative_to(self.active_folder)
            except ValueError:
                return f"Path escapes the active folder: {new_filename}"
            if dst.exists():
                return f"Target already exists: {new_name}"
            dst.parent.mkdir(parents=True, exist_ok=True)
            src.rename(dst)
            return json.dumps(
                {
                    "old": old_filename,
                    "new": new_name,
                    "full_path": str(dst),
                },
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Rename file error: {e}"

    def copy_local_excel_file(self, src_filename: str, dst_filename: str) -> str:
        """Duplicate a .xlsx file inside the active workspace.

        Args:
            src_filename: Existing file relative to active_folder.
            dst_filename: New file name relative to active_folder.

        Returns:
            JSON with paths or error.
        """
        if self.active_folder is None:
            return "No folder is open. Call open_local_folder(path) first."
        try:
            import shutil

            src = self._resolve_file(src_filename)
            new_name = (
                dst_filename if dst_filename.lower().endswith(".xlsx") else f"{dst_filename}.xlsx"
            )
            dst = (self.active_folder / new_name).resolve()
            try:
                dst.relative_to(self.active_folder)
            except ValueError:
                return f"Path escapes the active folder: {dst_filename}"
            if dst.exists():
                return f"Target already exists: {new_name}"
            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, dst)
            return json.dumps(
                {
                    "source": src_filename,
                    "copy": new_name,
                    "full_path": str(dst),
                },
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Copy file error: {e}"

    def move_local_excel_file(self, src_filename: str, dst_filename: str) -> str:
        """Move a .xlsx file to a different sub-path within the workspace.

        Both source and destination must stay inside the active folder.
        Differs from rename: `dst_filename` may include subdirectories.

        Args:
            src_filename: Existing file relative to active_folder.
            dst_filename: New location relative to active_folder (may include
                subdirs; missing dirs are created).

        Returns:
            JSON with paths or error.
        """
        if self.active_folder is None:
            return "No folder is open. Call open_local_folder(path) first."
        try:
            import shutil

            src = self._resolve_file(src_filename)
            new_name = (
                dst_filename if dst_filename.lower().endswith(".xlsx") else f"{dst_filename}.xlsx"
            )
            dst = (self.active_folder / new_name).resolve()
            try:
                dst.relative_to(self.active_folder)
            except ValueError:
                return f"Path escapes the active folder: {dst_filename}"
            if dst.exists():
                return f"Target already exists: {new_name}"
            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(src), str(dst))
            return json.dumps(
                {
                    "from": src_filename,
                    "to": new_name,
                    "full_path": str(dst),
                },
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Move file error: {e}"

    # ---------- Column operations + sort ----------

    def sort_local_excel_by_column(
        self,
        filename: str,
        sheet_name: str,
        column: str,
        ascending: bool = True,
        has_header: bool = True,
    ) -> str:
        """Sort rows by a column in a local .xlsx file (saved in place).

        Args:
            filename: Path relative to active_folder.
            sheet_name: Sheet to sort.
            column: Column letter ("B") or 1-based index as string ("2").
            ascending: True for ascending sort; False for descending.
            has_header: When True, the first row stays in place.

        Returns:
            Success message with row counts or error.
        """
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
            col_idx = _resolve_column(column)

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return f"{filename}!{sheet_name}: empty sheet, nothing to sort."

            header = rows[0] if has_header else None
            body = rows[1:] if has_header else rows

            def key(row: tuple) -> Any:
                v = row[col_idx - 1] if col_idx - 1 < len(row) else None
                if v is None:
                    return (1, "")  # Nones sort last
                if isinstance(v, (int, float)):
                    return (0, v)
                return (0, str(v).lower())

            body.sort(key=key, reverse=not ascending)

            # Rewrite the sheet
            ws.delete_rows(1, ws.max_row)
            if header is not None:
                ws.append(list(header))
            for r in body:
                ws.append(list(r))
            wb.save(path)
            wb.close()
            return (
                f"{filename}!{sheet_name}: sorted by column {column} "
                f"({'asc' if ascending else 'desc'}), {len(body)} rows (saved in place)"
            )
        except Exception as e:
            return f"Sort error: {e}"

    def delete_local_excel_columns(
        self, filename: str, sheet_name: str, start_col: str, count: int = 1
    ) -> str:
        """Delete one or more contiguous columns from a local .xlsx file.

        Args:
            filename: Path relative to active_folder.
            sheet_name: Sheet name.
            start_col: First column to delete (letter or 1-based index string).
            count: Number of columns to delete (default 1).

        Returns:
            Success message or error.
        """
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            col_idx = _resolve_column(start_col)
            wb[sheet_name].delete_cols(col_idx, count)
            wb.save(path)
            wb.close()
            return f"{filename}!{sheet_name}: {count} column(s) deleted from {start_col} (saved in place)"
        except Exception as e:
            return f"Delete columns error: {e}"

    def insert_local_excel_columns(
        self, filename: str, sheet_name: str, before_col: str, count: int = 1
    ) -> str:
        """Insert blank columns before the given column in a local .xlsx.

        Args:
            filename: Path relative to active_folder.
            sheet_name: Sheet name.
            before_col: Insertion anchor (letter or 1-based index string).
            count: Number of blank columns to insert.

        Returns:
            Success message or error.
        """
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path)
            if sheet_name not in wb.sheetnames:
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            col_idx = _resolve_column(before_col)
            wb[sheet_name].insert_cols(col_idx, count)
            wb.save(path)
            wb.close()
            return f"{filename}!{sheet_name}: {count} blank column(s) inserted before {before_col} (saved in place)"
        except Exception as e:
            return f"Insert columns error: {e}"

    # ---------- Analytical + export ----------

    def describe_local_excel(
        self, filename: str, sheet_name: Optional[str] = None, sample_size: int = 5
    ) -> str:
        """Produce a pandas-`describe`-style overview of a local .xlsx file.

        For each sheet (or just one if `sheet_name` given), reports per
        column: header value, inferred type, total/non-empty count,
        unique-value count, min/max for numeric columns, and a few sample
        values.

        Args:
            filename: Path relative to active_folder.
            sheet_name: Restrict to a single sheet. None -> describe all.
            sample_size: Number of sample values to include per column.

        Returns:
            JSON: {sheets: {sheet_name: {rows, columns: [...]}}, ...}.
        """
        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path, read_only=True, data_only=True)
            sheets = [sheet_name] if sheet_name else wb.sheetnames
            for s in sheets:
                if s not in wb.sheetnames:
                    wb.close()
                    return f"Sheet not found: {s}. Available: {wb.sheetnames}"

            result: dict[str, Any] = {"file": filename, "sheets": {}}
            for s in sheets:
                ws = wb[s]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    result["sheets"][s] = {"rows": 0, "columns": []}
                    continue

                header = rows[0]
                body = rows[1:]
                col_count = max(len(r) for r in rows) if rows else 0

                columns_info: list[dict[str, Any]] = []
                for col_idx in range(col_count):
                    values = [
                        r[col_idx] if col_idx < len(r) else None
                        for r in body
                    ]
                    non_empty = [
                        v for v in values
                        if v is not None and not (isinstance(v, str) and not v.strip())
                    ]
                    nums: list[float] = []
                    non_numeric = 0
                    for v in non_empty:
                        try:
                            nums.append(float(v))
                        except (TypeError, ValueError):
                            non_numeric += 1

                    inferred = (
                        "numeric"
                        if nums and non_numeric == 0
                        else "mixed" if nums and non_numeric > 0
                        else "text" if non_empty
                        else "empty"
                    )
                    samples = [str(v) for v in non_empty[:sample_size]]
                    info: dict[str, Any] = {
                        "header": str(header[col_idx]) if col_idx < len(header) else None,
                        "letter": get_column_letter(col_idx + 1),
                        "type": inferred,
                        "total": len(values),
                        "non_empty": len(non_empty),
                        "empty": len(values) - len(non_empty),
                        "unique": len({str(v) for v in non_empty}),
                        "samples": samples,
                    }
                    if nums:
                        info["min"] = min(nums)
                        info["max"] = max(nums)
                        info["mean"] = sum(nums) / len(nums)
                    columns_info.append(info)

                result["sheets"][s] = {"rows": len(body), "columns": columns_info}
            wb.close()
            return json.dumps(result, ensure_ascii=False, default=str)
        except Exception as e:
            return f"Describe error: {e}"

    def search_in_all_local_files(
        self,
        pattern: str,
        case_sensitive: bool = False,
        use_regex: bool = False,
        whole_cell: bool = False,
        recursive: bool = False,
        max_files: int = 100,
        max_hits_per_file: int = 50,
    ) -> str:
        """Search a pattern across every .xlsx file in the active workspace.

        Returns a JSON summary: per file, which sheets/cells matched.

        Args:
            pattern: Substring or regex.
            case_sensitive: Case sensitivity flag.
            use_regex: Treat pattern as regex.
            whole_cell: Require the entire cell to equal the pattern.
            recursive: Walk subdirectories too.
            max_files: Hard cap on files scanned.
            max_hits_per_file: Per-file cap on returned matches.

        Returns:
            JSON: {pattern, total_hits, files: [{file, sheets: {...}}]}.
        """
        if self.active_folder is None:
            return "No folder is open. Call open_local_folder(path) first."
        try:
            glob_pat = "**/*.xlsx" if recursive else "*.xlsx"
            files = [
                p
                for p in self.active_folder.glob(glob_pat)
                if p.is_file()
                and not p.name.startswith("~$")
                and not _is_ignored(p.relative_to(self.active_folder))
            ][:max_files]

            regex: Optional[re.Pattern[str]] = None
            if use_regex:
                flags = 0 if case_sensitive else re.IGNORECASE
                regex = re.compile(pattern, flags)
            needle = pattern if case_sensitive else pattern.lower()

            per_file: list[dict[str, Any]] = []
            total_hits = 0
            for fpath in files:
                rel = fpath.relative_to(self.active_folder).as_posix()
                file_entry: dict[str, Any] = {"file": rel, "sheets": {}}
                try:
                    wb = load_workbook(fpath, read_only=True, data_only=True)
                except Exception as inner:
                    file_entry["error"] = str(inner)
                    per_file.append(file_entry)
                    continue
                for s in wb.sheetnames:
                    ws = wb[s]
                    hits: list[dict[str, Any]] = []
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value is None:
                                continue
                            text = str(cell.value)
                            cmp_text = text if case_sensitive else text.lower()
                            if use_regex and regex:
                                m = bool(regex.fullmatch(text) if whole_cell else regex.search(text))
                            else:
                                m = (cmp_text == needle) if whole_cell else (needle in cmp_text)
                            if m:
                                hits.append({"cell": cell.coordinate, "value": text})
                                if len(hits) >= max_hits_per_file:
                                    break
                        if len(hits) >= max_hits_per_file:
                            break
                    if hits:
                        file_entry["sheets"][s] = hits
                        total_hits += len(hits)
                wb.close()
                if file_entry["sheets"] or file_entry.get("error"):
                    per_file.append(file_entry)

            return json.dumps(
                {
                    "pattern": pattern,
                    "files_scanned": len(files),
                    "files_with_matches": sum(1 for f in per_file if f.get("sheets")),
                    "total_hits": total_hits,
                    "files": per_file,
                },
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Search error: {e}"

    def export_local_excel_to_csv(
        self,
        filename: str,
        sheet_name: str,
        output_filename: Optional[str] = None,
    ) -> str:
        """Export a local .xlsx sheet to a CSV file in the same workspace.

        Args:
            filename: Source .xlsx path relative to active_folder.
            sheet_name: Sheet to export.
            output_filename: Output .csv path relative to active_folder.
                None -> derived from input (e.g. "data.xlsx" -> "data.csv").

        Returns:
            JSON: {source, output, rows} or error.
        """
        import csv

        try:
            path = self._resolve_file(filename)
            wb = load_workbook(path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}"
            ws = wb[sheet_name]

            assert self.active_folder is not None
            if output_filename is None:
                base = path.stem
                output_filename = f"{base}.csv"
            elif not output_filename.lower().endswith(".csv"):
                output_filename = f"{output_filename}.csv"

            out_path = (self.active_folder / output_filename).resolve()
            try:
                out_path.relative_to(self.active_folder)
            except ValueError:
                wb.close()
                return f"Path escapes the active folder: {output_filename}"
            out_path.parent.mkdir(parents=True, exist_ok=True)

            row_count = 0
            with out_path.open("w", encoding="utf-8", newline="") as fh:
                writer = csv.writer(fh)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(["" if v is None else v for v in row])
                    row_count += 1
            wb.close()
            return json.dumps(
                {
                    "source": filename,
                    "sheet": sheet_name,
                    "output": output_filename,
                    "rows": row_count,
                    "full_path": str(out_path),
                },
                ensure_ascii=False,
            )
        except Exception as e:
            return f"Export error: {e}"

    # ---------- Internal helpers ----------

    def _resolve_file(self, filename: str) -> Path:
        """Resolve a filename to a safe absolute Path inside active_folder."""
        if self.active_folder is None:
            raise RuntimeError(
                "No folder is open. Call open_local_folder(path) first."
            )
        candidate = (self.active_folder / filename).resolve()
        try:
            candidate.relative_to(self.active_folder)
        except ValueError:
            raise RuntimeError(
                f"Path escapes the active folder: {filename}"
            )
        if not candidate.exists():
            raise RuntimeError(f"File not found: {candidate}")
        return candidate

    def _resolve_subpath(self, subpath: str) -> Path:
        """Resolve a subpath (for listings) safely under active_folder."""
        assert self.active_folder is not None
        candidate = (self.active_folder / subpath).resolve()
        try:
            candidate.relative_to(self.active_folder)
        except ValueError:
            raise RuntimeError(f"Subpath escapes the active folder: {subpath}")
        return candidate

    def _ls(self, target: Path) -> list[dict[str, Any]]:
        """Return a sorted directory listing as a list of dicts."""
        entries: list[dict[str, Any]] = []
        try:
            items = sorted(target.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower()))
        except OSError as e:
            return [{"error": f"Cannot read directory: {e}"}]
        for item in items:
            try:
                stat = item.stat()
                entries.append(
                    {
                        "name": item.name,
                        "type": "dir" if item.is_dir() else "file",
                        "size": stat.st_size if item.is_file() else None,
                        "ext": item.suffix.lower() if item.is_file() else None,
                    }
                )
            except OSError:
                continue
        return entries

    def _walk(self, target: Path, max_results: int) -> list[dict[str, Any]]:
        """Recursively walk `target`, returning entries with relative paths.

        Uses a pruning BFS so we can skip noise directories (IGNORE_DIRS)
        instead of walking through them.
        """
        entries: list[dict[str, Any]] = []
        stack: list[Path] = [target]
        while stack and len(entries) < max_results:
            current = stack.pop(0)
            try:
                children = sorted(current.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower()))
            except OSError:
                continue
            for p in children:
                if p.is_dir() and p.name in IGNORE_DIRS:
                    continue
                try:
                    rel = p.relative_to(target)
                    stat = p.stat()
                    entries.append(
                        {
                            "name": rel.as_posix(),
                            "type": "dir" if p.is_dir() else "file",
                            "size": stat.st_size if p.is_file() else None,
                            "ext": p.suffix.lower() if p.is_file() else None,
                        }
                    )
                except OSError:
                    continue
                if len(entries) >= max_results:
                    break
                if p.is_dir():
                    stack.append(p)
        return entries
